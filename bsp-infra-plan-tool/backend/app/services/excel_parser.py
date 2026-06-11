import logging
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.models.models import parse_line_type
from app.services.report_types import (
    REQUIRED_COLUMNS,
    ParseResult,
    ParsedOrder,
    cell_text,
    finalize_parse_result,
    is_allowed_order_type,
    parse_date_value,
    parse_datetime_value,
)

logger = logging.getLogger(__name__)

INFO_SHEET = "Info"
ORDERS_SHEET = "Lijn orders"
REPORT_DATE_LABEL = "Aangemaakt op"


def _parse_info_report_date(workbook) -> datetime | None:
    if INFO_SHEET not in workbook.sheetnames:
        return None
    ws = workbook[INFO_SHEET]
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row and cell_text(row[0]) == REPORT_DATE_LABEL and len(row) > 1:
            return parse_datetime_value(row[1])
    return None


def _header_map(ws) -> dict[str, int]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError(f"Tabblad '{ORDERS_SHEET}' bevat geen kolomkoppen")
    column_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = cell_text(cell)
        if name:
            column_map[name] = idx
    return column_map


def parse_excel_xlsx(content: bytes) -> ParseResult:
    result = ParseResult()
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Ongeldig Excel-bestand: {exc}") from exc

    try:
        if ORDERS_SHEET not in workbook.sheetnames:
            raise ValueError(f"Tabblad '{ORDERS_SHEET}' niet gevonden in het bestand")

        result.report_date = _parse_info_report_date(workbook)
        ws = workbook[ORDERS_SHEET]
        column_map = _header_map(ws)

        missing = REQUIRED_COLUMNS - set(column_map.keys())
        if missing:
            raise ValueError(f"Ontbrekende kolommen: {', '.join(sorted(missing))}")

        def get_col(row: tuple, name: str) -> str:
            idx = column_map.get(name)
            if idx is None or idx >= len(row):
                return ""
            return cell_text(row[idx])

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None or cell_text(v) == "" for v in row):
                continue

            order_number = get_col(row, "Order number")
            bedrijf = get_col(row, "Bedrijf")
            geplaatst_raw = row[column_map["Geplaatst op"]] if "Geplaatst op" in column_map else None
            gepland_raw = row[column_map["Gepland"]] if "Gepland" in column_map else None
            line_type_raw = get_col(row, "On-/Offnet")

            if not order_number:
                result.skipped_rows += 1
                continue

            order_type = get_col(row, "Order type")
            if not is_allowed_order_type(order_type):
                result.filtered_order_type += 1
                continue

            geplaatst_op = parse_date_value(geplaatst_raw)
            gepland = parse_date_value(gepland_raw)
            line_type = parse_line_type(line_type_raw)

            if not bedrijf or not geplaatst_op or not gepland or not line_type:
                result.warnings.append(
                    f"Rij {row_idx} overgeslagen: ontbrekende verplichte velden voor order {order_number}"
                )
                result.skipped_rows += 1
                continue

            result.orders.append(
                ParsedOrder(
                    order_number=order_number,
                    bedrijf=bedrijf,
                    geplaatst_op=geplaatst_op,
                    gepland=gepland,
                    line_type=line_type.value,
                )
            )
    finally:
        workbook.close()

    logger.info(
        "Parsed %d orders from Excel, skipped %d rows, filtered %d by order type",
        len(result.orders),
        result.skipped_rows,
        result.filtered_order_type,
    )
    return finalize_parse_result(result)
